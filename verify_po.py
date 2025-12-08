import requests
import os
import glob
import openpyxl
from openpyxl.utils import get_column_letter

def verify_po():
    # 1. Trigger endpoint
    url = "http://localhost:8000/generate-production-order/a34fD000001KmeI"
    print(f"Calling {url}...")
    try:
        response = requests.get(url)
        response.raise_for_status()
        print("Endpoint called successfully.")
    except Exception as e:
        print(f"Error calling endpoint: {e}")
        return

    # 2. Find generated file
    output_dir = "output"
    files = glob.glob(os.path.join(output_dir, "*.xlsx"))
    if not files:
        print("No Excel files found in output directory.")
        return
    
    # Get latest file
    latest_file = max(files, key=os.path.getctime)
    print(f"Checking file: {latest_file}")
    
    wb = openpyxl.load_workbook(latest_file)
    ws = wb.active
    
    # 3. Verify Totals (Formulas)
    # Assuming totals are in columns H-M (8-13)
    # Find Total row
    total_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=4).value
        if val and "TỔNG CỘNG" in str(val).upper():
            total_row = r
            break
            
    if not total_row:
        print("FAIL: 'TỔNG CỘNG' row not found.")
    else:
        print(f"Found 'TỔNG CỘNG' at row {total_row}")
        
        # Check formulas
        cols_to_check = [8, 9, 10, 11, 12, 13]
        all_formulas = True
        for col in cols_to_check:
            cell = ws.cell(row=total_row, column=col)
            val = cell.value
            if not isinstance(val, str) or not val.startswith("=SUM"):
                print(f"FAIL: Column {get_column_letter(col)} does not have SUM formula. Value: {val}")
                all_formulas = False
            else:
                print(f"PASS: Column {get_column_letter(col)} has formula: {val}")
        
        if all_formulas:
            print("PASS: All total columns have SUM formulas.")

        # Check unmerged
        is_merged = False
        for col in cols_to_check:
            cell = ws.cell(row=total_row, column=col)
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    print(f"FAIL: Cell {cell.coordinate} is merged.")
                    break
        if not is_merged:
            print("PASS: Total cells are not merged.")

        # Check formatting (Bold, Center, Border)
        # Just checking one cell as sample
        cell = ws.cell(row=total_row, column=8)
        if cell.font.bold:
            print("PASS: Total row is Bold.")
        else:
            print("FAIL: Total row is NOT Bold.")
            
        if cell.alignment.horizontal == 'center':
            print("PASS: Total row is Center aligned.")
        else:
            print("FAIL: Total row is NOT Center aligned.")
            
        if cell.border.left.style == 'thin': # Simple check
            print("PASS: Total row has Border.")
        else:
            print("FAIL: Total row might not have Border.")

    # 4. Verify "Ngọc Bích" and Merge I-K
    signer_found = False
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=9).value
        if val == "Ngọc Bích":
            signer_found = True
            print(f"PASS: Found 'Ngọc Bích' at row {r}")
            # Check merge
            cell = ws.cell(row=r, column=9)
            is_merged_correctly = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    if merged_range.min_col == 9 and merged_range.max_col == 11: # I to K
                        is_merged_correctly = True
                        print(f"PASS: 'Ngọc Bích' cell is merged I-K ({merged_range})")
                    else:
                        print(f"FAIL: 'Ngọc Bích' cell merged range is {merged_range}, expected I-K")
                    break
            if not is_merged_correctly:
                print("FAIL: 'Ngọc Bích' cell is NOT merged correctly.")
            break
            
    if not signer_found:
        print("FAIL: 'Ngọc Bích' not found.")

    # 5. Verify Product Name and Delivery Date Merging
    # We can check if there are merged cells in col 4 and 15
    # This is a bit harder to verify automatically without knowing expected data, 
    # but we can check if there are ANY merged cells in those columns within the data range.
    print("Checking merges in Product Name (D) and Delivery Date (O)...")
    # Assuming data starts around row 13 (or where table starts)
    # We can just list merged ranges in those columns
    merged_D = []
    merged_O = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col == 4 and merged_range.max_col == 4:
            merged_D.append(str(merged_range))
        if merged_range.min_col == 15 and merged_range.max_col == 15:
            merged_O.append(str(merged_range))
            
    print(f"Merged ranges in Column D: {merged_D}")
    print(f"Merged ranges in Column O: {merged_O}")
    if merged_D:
        print("PASS: Found merged cells in Product Name column (indicates merging logic ran).")
    else:
        print("WARNING: No merged cells in Product Name column. (Might be correct if no duplicates)")
        
    if merged_O:
        print("PASS: Found merged cells in Delivery Date column (indicates merging logic ran).")
    else:
        print("WARNING: No merged cells in Delivery Date column. (Might be correct if no duplicates)")

if __name__ == "__main__":
    verify_po()
