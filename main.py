from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from simple_salesforce import Salesforce
from dotenv import load_dotenv
import openpyxl
from copy import copy as style_copy
from openpyxl.utils import get_column_letter
import base64
import datetime
import os
import json
import re
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from pathlib import Path
from openpyxl.styles import Alignment

# Load environment variables
load_dotenv()

app = FastAPI(title="Salesforce Packing List API")

class ShipmentRequest(BaseModel):
    shipment_id: str

def expand_items_table(ws, template_row, n):
    """Expand the items table to accommodate n rows"""
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=template_row, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    row_height = ws.row_dimensions[template_row].height
    add_rows = max(0, n - 1)
    
    # Handle merged cells
    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > template_row:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
    
    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)
    
    # Insert rows
    if add_rows > 0:
        ws.insert_rows(template_row + 1, amount=add_rows)
        for offset in range(1, add_rows + 1):
            r = template_row + offset
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                dst.value = None
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
            if row_height is not None:
                ws.row_dimensions[r].height = row_height
    
    # Re-merge shifted cells
    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)
    
    # Update total formulas
    total_header_row = None
    for r in range(template_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Total":
            total_header_row = r
            break
    
    if total_header_row is None:
        raise ValueError("Total row not found")
    
    first_data_row = template_row
    last_data_row = template_row + n - 1
    ws[f"H{total_header_row}"] = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws[f"J{total_header_row}"] = f"=SUM(J{first_data_row}:J{last_data_row})"
    ws[f"K{total_header_row}"] = f"=COUNTA(K{first_data_row}:K{last_data_row})"

def get_salesforce_connection():
    """Initialize Salesforce connection"""
    username = os.getenv('SALESFORCE_USERNAME')
    password = os.getenv('SALESFORCE_PASSWORD')
    security_token = os.getenv('SALESFORCE_SECURITY_TOKEN')
    consumer_key = os.getenv('SALESFORCE_CONSUMER_KEY')
    consumer_secret = os.getenv('SALESFORCE_CONSUMER_SECRET')
    
    if None in (username, password, security_token, consumer_key, consumer_secret):
        raise ValueError("Salesforce credentials missing in environment variables")
    
    return Salesforce(
        username=username,
        password=password,
        security_token=security_token,
        consumer_key=consumer_key,
        consumer_secret=consumer_secret
    )

def get_picklist_values(sf, object_name: str, field_name: str) -> list[str]:
    """
    Get picklist values dynamically from Salesforce for any object and field.
    
    Args:
        sf: Salesforce connection instance
        object_name: API name of the Salesforce object (e.g., 'Shipment__c')
        field_name: API name of the picklist field (e.g., 'Freight__c')
    
    Returns:
        List of picklist option values
    """
    try:
        sobject = getattr(sf, object_name)
        description = sobject.describe()
        
        for field in description['fields']:
            if field['name'] == field_name:
                if field['type'] == 'picklist' or field['type'] == 'multipicklist':
                    return [option['value'] for option in field['picklistValues'] if option['active']]
                else:
                    print(f"⚠ Warning: {field_name} is not a picklist field (type: {field['type']})")
                    return []
        
        print(f"⚠ Warning: {field_name} field not found on {object_name}")
        return []
    except Exception as e:
        print(f"⚠ Warning: Could not fetch picklist values for {object_name}.{field_name}: {e}")
        return []

def get_output_directory() -> Path:
    """
    Get the appropriate output directory based on environment.
    Use /tmp for serverless environments (Vercel, AWS Lambda) where filesystem is read-only.
    Use ./output for local development.
    """
    # Check if we're in a serverless environment
    is_serverless = (
        os.getenv('VERCEL') is not None or  # Vercel
        os.getenv('AWS_LAMBDA_FUNCTION_NAME') is not None or  # AWS Lambda
        os.getenv('LAMBDA_TASK_ROOT') is not None  # AWS Lambda alternative
    )
    
    if is_serverless:
        output_dir = Path("/tmp")
    else:
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
    
    return output_dir

def format_picklist_checkboxes(options: list[str], selected_value: str, uppercase: bool = False) -> str:
    """
    Format picklist options as checkbox text.
    
    Args:
        options: List of picklist options
        selected_value: The currently selected value
        uppercase: Whether to display values in uppercase
    
    Returns:
        Formatted string with checkboxes
    """
    checked_box = '☑'
    unchecked_box = '☐'
    
    selected_upper = (selected_value or '').strip().upper()
    
    lines = []
    for opt in options:
        mark = checked_box if opt.upper() == selected_upper else unchecked_box
        display_value = opt.upper() if uppercase else opt
        lines.append(f"{mark} {display_value}")
    
    return "\n".join(lines)

def generate_packing_list(shipment_id: str, template_path: str):
    """Generate packing list for a given shipment ID"""
    
    # Connect to Salesforce
    sf = get_salesforce_connection()
    
    # Get freight options dynamically from Salesforce
    freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
    
    # Query shipment data
    shipment_query = f"""
    SELECT Name, Consignee__c, Invoice_Packing_list_no__c, Issued_date__c, Port_of_Origin__c,
    Final_Destination__c, Stockyard__c, Ocean_Vessel__c, B_L_No__c, Freight__c,
    Departure_Date_ETD__c, Arrival_Schedule_ETA__c, Remark_number_on_documents__c,
    Terms_of_Sales__c, Terms_of_Payment__c
    FROM Shipment__c
    WHERE Id = '{shipment_id}'
    """
    shipment_result = sf.query(shipment_query)
    if not shipment_result['records']:
        raise ValueError(f"No Shipment found with ID: {shipment_id}")
    shipment = shipment_result['records'][0]
    
    # Query account/consignee data
    if shipment['Consignee__c']:
        account_query = f"""
        SELECT Name, BillingStreet, BillingCity, BillingPostalCode, BillingCountry,
        Phone, Fax__c, VAT__c
        FROM Account
        WHERE Id = '{shipment['Consignee__c']}'
        """
        account_result = sf.query(account_query)
        account = account_result['records'][0] if account_result['records'] else {}
    else:
        account = {}
    
    # Query bookings
    bookings_query = f"""
    SELECT Id, Cont_Quantity__c
    FROM Booking__c
    WHERE Shipment__c = '{shipment_id}'
    """
    bookings_result = sf.query_all(bookings_query)
    bookings = bookings_result['records']
    total_containers_from_bookings = sum(booking.get('Cont_Quantity__c') or 0 for booking in bookings)
    
    # Query container items
    items_query = f"""
    SELECT Line_item_no_for_print__c, Product_Description__c, Length__c, Width__c, Height__c,
    Quantity_For_print__c, Unit_for_print__c, Crates__c, Packing__c, Order_No__c,
    Container__r.Name, Container__r.Container_Weight_Regulation__c
    FROM Container_Item__c
    WHERE Shipment__c = '{shipment_id}'
    """
    items_result = sf.query_all(items_query)
    items = items_result['records']
    
    # Load template
    wb = openpyxl.load_workbook(template_path)
    ws = wb['PackingList']
    
    # Replace placeholders (excluding Freight__c as it needs special handling)
    replacements = {
        '{{Shipment__c.Consignee__r.Name}}': account.get('Name') or '',
        '{{Shipment__c.Consignee__r.BillingStreet}}': account.get('BillingStreet') or '',
        '{{Shipment__c.Consignee__r.BillingCity}}': account.get('BillingCity') or '',
        '{{Shipment__c.Consignee__r.BillingPostalCode}}': account.get('BillingPostalCode') or '',
        '{{Shipment__c.Consignee__r.BillingCountry}}': account.get('BillingCountry') or '',
        '{{Shipment__c.Consignee__r.Phone}}': account.get('Phone') or '',
        '{{Shipment__c.Consignee__r.Fax__c}}': account.get('Fax__c') or '',
        '{{Shipment__c.Consignee__r.VAT__c}}': account.get('VAT__c') or '',
        '{{Shipment__c.Invoice_Packing_list_no__c}}': shipment.get('Invoice_Packing_list_no__c') or '',
        '{{Shipment__c.Issued_date__c}}': shipment.get('Issued_date__c') or '',
        '{{Shipment__c.Port_of_Origin__c}}': shipment.get('Port_of_Origin__c') or '',
        '{{Shipment__c.Final_Destination__c}}': shipment.get('Final_Destination__c') or '',
        '{{Shipment__c.Stockyard__c}}': shipment.get('Stockyard__c') or '',
        '{{Shipment__c.Ocean_Vessel__c}}': shipment.get('Ocean_Vessel__c') or '',
        '{{Shipment__c.B_L_No__c}}': shipment.get('B_L_No__c') or '',
        '{{Shipment__c.Departure_Date_ETD__c}}': shipment.get('Departure_Date_ETD__c') or '',
        '{{Shipment__c.Arrival_Schedule_ETA__c}}': shipment.get('Arrival_Schedule_ETA__c') or '',
        '{{Shipment__c.Remark_number_on_documents__c}}': shipment.get('Remark_number_on_documents__c') or '',
        '{{Shipment__c.Terms_of_Sales__c}}': shipment.get('Terms_of_Sales__c') or '',
        '{{Shipment__c.Terms_of_Payment__c}}': shipment.get('Terms_of_Payment__c') or '',
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for placeholder, value in replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))
                if '{{TableStart:Shipment__c.r.Bookings__r}}' in cell.value:
                    cell.value = str(total_containers_from_bookings)
    
    # Remove "None" values
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'None' in cell.value:
                cell.value = cell.value.replace('None', '')
    
    # Handle freight checkboxes dynamically
    checked_box = '☑'
    unchecked_box = '☐'
    
    freight_value = (shipment.get('Freight__c') or '').strip()
    freight_upper = freight_value.upper()
    
    # Generate checkbox text with all options from Salesforce
    lines = []
    for opt in freight_options:
        mark = checked_box if opt.upper() == freight_upper else unchecked_box
        lines.append(f"{mark} {opt}")
    
    checkbox_text = "\n".join(lines)
    
    # Replace freight placeholder with checkbox text
    # for row in ws.iter_rows():
    #     for cell in row:
    #         if isinstance(cell.value, str) and '{{Shipment__c.Freight__c}}' in cell.value:
    #             cell.value = cell.value.replace('{{Shipment__c.Freight__c}}', checkbox_text)
    
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '{{Shipment__c.Freight__c}}' in cell.value:
                cell.value = cell.value.replace('{{Shipment__c.Freight__c}}', checkbox_text)
                if cell.alignment:
                    new_alignment = style_copy(cell.alignment)
                else:
                    from openpyxl.styles import Alignment
                    new_alignment = Alignment()
                new_alignment.wrap_text = True
                cell.alignment = new_alignment
    
    # Find table start row
    table_start_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            if cell.value and '{{TableStart:ContainerItems}}' in str(cell.value):
                table_start_row = cell.row
                break
        if table_start_row:
            break
    
    if not table_start_row:
        raise ValueError("No table start marker found in template")
    
    # Expand table
    expand_items_table(ws, table_start_row, len(items) if items else 1)
    
    # Fill in item data
    for idx, item in enumerate(items):
        row = table_start_row + idx
        container_r = item.get('Container__r', {})
        line_item_no = item.get('Line_item_no_for_print__c') or str(idx + 1)
        ws.cell(row, 1).value = line_item_no
        ws.cell(row, 2).value = item.get('Product_Description__c')
        ws.cell(row, 3).value = item.get('Length__c')
        ws.cell(row, 4).value = item.get('Width__c')
        ws.cell(row, 5).value = item.get('Height__c')
        ws.cell(row, 6).value = item.get('Quantity_For_print__c') or ''
        ws.cell(row, 7).value = item.get('Unit_for_print__c') or ''
        ws.cell(row, 8).value = item.get('Crates__c')
        ws.cell(row, 9).value = f"{item.get('Packing__c') or ''} pcs/crate"
        ws.cell(row, 10).value = container_r.get('Container_Weight_Regulation__c')
        ws.cell(row, 11).value = container_r.get('Name')
        ws.cell(row, 13).value = item.get('Order_No__c')
    
    # Save file
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"Packing_List_{shipment.get('Invoice_Packing_list_no__c', shipment['Name'])}_{timestamp}.xlsx"
    
    # Use appropriate output directory based on environment
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": shipment_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version['id'],
        "freight_options_used": freight_options
    }

@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "message": "Salesforce Packing List API",
        "version": "1.0.0",
        "endpoints": {
            "GET /health": "Health check",
            "GET /generate-packing-list": "Generate packing list (test endpoint)",
            "POST /generate-packing-list": "Generate packing list (production endpoint)",
            "GET /generate_invoice/{shipment_id}": "Generate invoice for a shipment",
            "GET /generate-combined-export/{shipment_id}": "Generate combined packing list and invoice in one Excel file",
            "GET /download/{file_name}": "Download generated packing list file"
        }
    }

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    try:
        # Test Salesforce connection
        sf = get_salesforce_connection()
        freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
        return {
            "status": "healthy",
            "salesforce_connected": True,
            "freight_options": freight_options,
            "timestamp": datetime.datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Service unhealthy: {str(e)}")

@app.get("/generate-packing-list")
async def generate_packing_list_get(shipment_id: str):
    """
    Generate packing list for a shipment (GET method for testing)
    
    Parameters:
    - shipment_id: Salesforce Shipment ID
    """
    try:
        template_path = os.getenv('TEMPLATE_PATH', 'templates/packing_list_template.xlsx')
        
        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=404,
                detail=f"Template file not found at: {template_path}"
            )
        
        result = generate_packing_list(shipment_id, template_path)
        
        return {
            "status": "success",
            "message": "Packing list generated successfully",
            "data": result
        }
    
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating packing list: {str(e)}")

@app.post("/generate-packing-list")
async def generate_packing_list_post(request: ShipmentRequest):
    """
    Generate packing list for a shipment (POST method)
    
    Parameters:
    - shipment_id: Salesforce Shipment ID (in request body)
    """
    try:
        template_path = os.getenv('TEMPLATE_PATH', 'templates/packing_list_template.xlsx')
        
        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=404,
                detail=f"Template file not found at: {template_path}"
            )
        
        result = generate_packing_list(request.shipment_id, template_path)
        
        return {
            "status": "success",
            "message": "Packing list generated successfully",
            "data": result
        }
    
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating packing list: {str(e)}")

def expand_invoice_items_table(ws, template_row: int, n: int) -> None:
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=template_row, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    row_height = ws.row_dimensions[template_row].height
    add_rows = max(0, n - 1)

    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > template_row:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))

    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)

    if add_rows > 0:
        ws.insert_rows(template_row + 1, amount=add_rows)
        for offset in range(1, add_rows + 1):
            r = template_row + offset
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                dst.value = None
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
            if row_height is not None:
                ws.row_dimensions[r].height = row_height

    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)

@app.get("/generate_invoice/{shipment_id}")
def generate_invoice(shipment_id: str):
    sf = get_salesforce_connection()

    # Base and discount templates
    base_template_path = "./templates/invoice_template.xlsx"
    discount_template_path = "./templates/invoice_template_w_discount.xlsx"
    template_path = base_template_path

    # Get all picklist values dynamically
    freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
    terms_of_sales_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Sales__c')
    terms_of_payment_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Payment__c')

    shipment_query = f"""
    SELECT Name, Consignee__c, Invoice_Packing_list_no__c, Issued_date__c,
           Port_of_Origin__c, Final_Destination__c, Stockyard__c,
           Ocean_Vessel__c, B_L_No__c, Freight__c,
           Departure_Date_ETD__c, Arrival_Schedule_ETA__c,
           Remark_number_on_documents__c,
           Terms_of_Sales__c, Terms_of_Payment__c,
           Subtotal_USD__c, Fumigation__c, In_words__c,
           Total_Price_USD__c, Surcharge_amount_USD__c,
           Discount_Percentage__c, Discount_Amount__c
    FROM Shipment__c
    WHERE Id = '{shipment_id}'
    """
    shipment_result = sf.query(shipment_query)
    if not shipment_result["records"]:
        raise ValueError(f"No Shipment found with ID: {shipment_id}")
    shipment = shipment_result["records"][0]

    # Determine if discount exists on the shipment
    discount_percentage = shipment.get("Discount_Percentage__c")
    discount_amount = shipment.get("Discount_Amount__c")

    discount_exists = any(
        v not in (None, 0, "", "0", 0.0)
        for v in (discount_percentage, discount_amount)
    )

    # Choose template based on discount
    if discount_exists:
        template_path = discount_template_path

    # Account / Consignee
    if shipment.get("Consignee__c"):
        account_query = f"""
        SELECT Name, BillingStreet, BillingCity, BillingPostalCode, BillingCountry,
               Phone, Fax__c, VAT__c
        FROM Account
        WHERE Id = '{shipment['Consignee__c']}'
        """
        account_result = sf.query(account_query)
        account = account_result["records"][0] if account_result["records"] else {}
    else:
        account = {}

    # Container items
    items_query = f"""
    SELECT Line_item_no_for_print__c, Product_Description__c,
           Length__c, Width__c, Height__c,
           Quantity_For_print__c, Unit_for_print__c,
           Sales_Price_USD__c, Charge_Unit__c,
           Total_Price_USD__c, Order_No__c,
           Container__r.STT_Cont__c
    FROM Container_Item__c
    WHERE Shipment__c = '{shipment_id}'
    ORDER BY Line_item_no_for_print__c
    """
    items_result = sf.query_all(items_query)
    items = items_result["records"]

    # 💡 FIXED: use existing fields only
    deposit_query = f"""
    SELECT Contract_PI__r.Name, Reconciled_Amount__c
    FROM Receipt_Reconciliation__c
    WHERE Invoice__c = '{shipment_id}'
    """
    deposits = sf.query_all(deposit_query)["records"]

    refunds_query = f"""
    SELECT Reason, Refund_Amount__c
    FROM Case
    WHERE Refund_in_Shipment__c = '{shipment_id}'
    """
    refunds = sf.query_all(refunds_query)["records"]

    # Build debug data for response
    debug_data = {
        "shipment": {k: v for k, v in shipment.items() if k != "attributes"},
        "account": {k: v for k, v in account.items() if k != "attributes"} if account else {},
        "container_items": [
            {k: v for k, v in item.items() if k != "attributes"}
            for item in items
        ],
        "deposits": [
            {k: v for k, v in dep.items() if k != "attributes"}
            for dep in deposits
        ],
        "refunds": [
            {k: v for k, v in ref.items() if k != "attributes"}
            for ref in refunds
        ],
        "picklist_options": {
            "Freight__c": freight_options,
            "Terms_of_Sales__c": terms_of_sales_options,
            "Terms_of_Payment__c": terms_of_payment_options,
        },
        "discount_exists": discount_exists,
        "template_used": template_path,
    }

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Invoice"] if "Invoice" in wb.sheetnames else wb.active

    # Format Port of Origin in uppercase
    port_of_origin = (shipment.get("Port_of_Origin__c") or "").upper()

    replacements = {
        "{{Shipment__c.Consignee__r.Name}}": account.get("Name") or "",
        "{{Shipment__c.Consignee__r.BillingStreet}}": account.get("BillingStreet") or "",
        "{{Shipment__c.Consignee__r.BillingCity}}": account.get("BillingCity") or "",
        "{{Shipment__c.Consignee__r.BillingPostalCode}}": account.get("BillingPostalCode") or "",
        "{{Shipment__c.Consignee__r.BillingCountry}}": account.get("BillingCountry") or "",
        "{{Shipment__c.Consignee__r.Phone}}": account.get("Phone") or "",
        "{{Shipment__c.Consignee__r.Fax__c}}": account.get("Fax__c") or "",
        "{{Shipment__c.Consignee__r.VAT__c}}": account.get("VAT__c") or "",
        "{{Shipment__c.Invoice_Packing_list_no__c}}": shipment.get("Invoice_Packing_list_no__c") or "",
        "{{Shipment__c.Issued_date__c}}": shipment.get("Issued_date__c") or "",
        "{{Shipment__c.Port_of_Origin__c}}": port_of_origin,
        "{{Shipment__c.Final_Destination__c}}": shipment.get("Final_Destination__c") or "",
        "{{Shipment__c.Stockyard__c}}": shipment.get("Stockyard__c") or "",
        "{{Shipment__c.Ocean_Vessel__c}}": shipment.get("Ocean_Vessel__c") or "",
        "{{Shipment__c.B_L_No__c}}": shipment.get("B_L_No__c") or "",
        "{{Shipment__c.Departure_Date_ETD__c}}": shipment.get("Departure_Date_ETD__c") or "",
        "{{Shipment__c.Arrival_Schedule_ETA__c}}": shipment.get("Arrival_Schedule_ETA__c") or "",
        "{{Shipment__c.Remark_number_on_documents__c}}": shipment.get("Remark_number_on_documents__c") or "",
        "{{Shipment__c.Subtotal_USD__c\\# #,##0.##}}": shipment.get("Subtotal_USD__c") or 0,
        "{{Shipment__c.Fumigation__c}}": shipment.get("Fumigation__c") or "",
        "{{Shipment__c.Total_Price_USD__c\\# #,##0.##}}": shipment.get("Total_Price_USD__c") or 0,
        "{{Shipment__c.In_words__c}}": shipment.get("In_words__c") or "",

        # 🔹 NEW: discount placeholders used by invoice_template_w_discount.xlsx
        "{{Shipment__c.Discount_Percentage__c}}": shipment.get("Discount_Percentage__c") or "",
        "{{Shipment__c.Discount_Amount__c\\# #,##0.##}}": shipment.get("Discount_Amount__c") or 0,
    }

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for placeholder, value in replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))

    # Clean "None"
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "None" in cell.value:
                cell.value = cell.value.replace("None", "")

    # 💡 FIXED: correct parameter order for checkboxes + uppercase
    freight_checkbox_text = format_picklist_checkboxes(
        freight_options, shipment.get("Freight__c"), uppercase=True
    )
    terms_of_sales_checkbox_text = format_picklist_checkboxes(
        terms_of_sales_options, shipment.get("Terms_of_Sales__c"), uppercase=True
    )
    terms_of_payment_checkbox_text = format_picklist_checkboxes(
        terms_of_payment_options, shipment.get("Terms_of_Payment__c"), uppercase=True
    )

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Shipment__c.Freight__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Freight__c}}", freight_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Sales__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Sales__c}}", terms_of_sales_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Payment__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)

    # --- ContainerItems table expansion ---
    table_start_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and "{{TableStart:ContainerItems}}" in cell.value:
                table_start_row = cell.row
                break
        if table_start_row:
            break

    if not table_start_row:
        raise ValueError("No ContainerItems table start marker found in template")

    expand_invoice_items_table(ws, table_start_row, len(items) if items else 1)

    for idx, item in enumerate(items):
        row_idx = table_start_row + idx
        container_r = item.get("Container__r") or {}
        line_item_no = item.get("Line_item_no_for_print__c") or str(idx + 1)
        ws.cell(row_idx, 1).value = line_item_no
        ws.cell(row_idx, 2).value = item.get("Product_Description__c")
        ws.cell(row_idx, 3).value = item.get("Length__c")
        ws.cell(row_idx, 4).value = item.get("Width__c")
        ws.cell(row_idx, 5).value = item.get("Height__c")
        ws.cell(row_idx, 6).value = item.get("Quantity_For_print__c")
        ws.cell(row_idx, 7).value = item.get("Unit_for_print__c")
        ws.cell(row_idx, 8).value = container_r.get("STT_Cont__c") or container_r.get("Name")
        ws.cell(row_idx, 9).value = f"{item.get('Sales_Price_USD__c') or ''} {item.get('Charge_Unit__c') or ''}".strip()
        ws.cell(row_idx, 10).value = item.get("Total_Price_USD__c")
        ws.cell(row_idx, 11).value = item.get("Order_No__c")

    # --- Deposits / refunds / surcharge sections (back to working behaviour) ---
    deposit_text_cell = None
    deposit_amount_cell = None
    refund_cell = None
    surcharge_text_cell = None
    surcharge_amount_cell = None

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            val = cell.value
            if "{{TableStart:InvoiceDeposit}}" in val:
                deposit_text_cell = cell
            if "Reconciled_Amount__c" in val:
                deposit_amount_cell = cell
            if "{{TableStart:Shipment__c.r.Cases__r}}" in val:
                refund_cell = cell
            if "{{TableStart:Surcharges}}" in val:
                surcharge_text_cell = cell
            if "Surcharge_amount_USD__c" in val:
                surcharge_amount_cell = cell

    # Deposits: multi-line "Deduct: Deposit of PI X"
    if deposit_text_cell and deposit_amount_cell:
        if deposits:
            labels = []
            amounts = []
            for rec in deposits:
                pi_name = (rec.get("Contract_PI__r") or {}).get("Name") or ""
                labels.append(f"Deduct: Deposit of PI {pi_name}".strip())
                amt = rec.get("Reconciled_Amount__c")
                amounts.append("" if amt is None else f"{amt:,.2f}")
            deposit_text_cell.value = "\n".join(labels)
            deposit_amount_cell.value = "\n".join(amounts)
        else:
            deposit_text_cell.value = None
            deposit_amount_cell.value = None

    # Refunds section
    if refund_cell:
        if refunds:
            lines = []
            for rec in refunds:
                reason = rec.get("Reason") or ""
                amt = rec.get("Refund_Amount__c")
                part = reason
                if amt is not None:
                    part = f"{reason} {amt:,.2f}".strip()
                lines.append(part)
            refund_cell.value = "\n".join(lines)
        else:
            refund_cell.value = None

    # Clean remaining refund placeholders if no refunds
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Refund_Amount__c" in cell.value or "{{TableEnd:Shipment__c.r.Cases__r}}" in cell.value:
                    if not refunds:
                        cell.value = cell.value.replace("{{Refund_Amount__c\\# #,##0.##}}", "")
                        cell.value = cell.value.replace("{{TableEnd:Shipment__c.r.Cases__r}}", "")
                        if cell.value and not cell.value.strip():
                            cell.value = None

    # Surcharge
    surcharge_amount = shipment.get("Surcharge_amount_USD__c")
    if surcharge_text_cell or surcharge_amount_cell:
        if surcharge_amount:
            if surcharge_text_cell:
                surcharge_text_cell.value = "Surcharge:"
            if surcharge_amount_cell:
                surcharge_amount_cell.value = f"{surcharge_amount:,.2f}"
        else:
            if surcharge_text_cell:
                surcharge_text_cell.value = None
            if surcharge_amount_cell:
                surcharge_amount_cell.value = None


    # Save file
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"Invoice_{shipment.get('Invoice_Packing_list_no__c', shipment['Name'])}_{timestamp}.xlsx"

    # Use appropriate output directory based on environment
    output_dir = get_output_directory()
    file_path = output_dir / file_name

    wb.save(str(file_path))

    # Upload to Salesforce as ContentVersion
    with open(file_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode("utf-8")

    content_version = sf.ContentVersion.create(
        {
            "Title": file_name.rsplit(".", 1)[0],
            "PathOnClient": file_name,
            "VersionData": encoded,
            "FirstPublishLocationId": shipment_id,
        }
    )

    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"],
        "freight_options_used": freight_options,
        "deposit_count": len(deposits),
        "refund_count": len(refunds),
        "discount_exists": discount_exists,
        "template_used": template_path,
        "debug_data": debug_data,
    }

@app.get("/generate-combined-export/{shipment_id}")
def generate_combined_export(shipment_id: str):
    """
    Generate combined packing list and invoice in one Excel file with two sheets.
    First sheet: "Packing List"
    Second sheet: "Invoice"
    
    Parameters:
    - shipment_id: Salesforce Shipment ID
    """
    sf = get_salesforce_connection()
    
    # Templates
    packing_list_template_path = os.getenv('TEMPLATE_PATH', 'templates/packing_list_template.xlsx')
    base_invoice_template_path = "./templates/invoice_template.xlsx"
    discount_invoice_template_path = "./templates/invoice_template_w_discount.xlsx"
    
    # Verify templates exist
    if not os.path.exists(packing_list_template_path):
        raise HTTPException(
            status_code=404,
            detail=f"Packing list template not found at: {packing_list_template_path}"
        )
    
    # Get picklist values dynamically
    freight_options = get_picklist_values(sf, 'Shipment__c', 'Freight__c')
    terms_of_sales_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Sales__c')
    terms_of_payment_options = get_picklist_values(sf, 'Shipment__c', 'Terms_of_Payment__c')
    
    # Query shipment data (combining fields from both packing list and invoice)
    shipment_query = f"""
    SELECT Name, Consignee__c, Invoice_Packing_list_no__c, Issued_date__c,
           Port_of_Origin__c, Final_Destination__c, Stockyard__c,
           Ocean_Vessel__c, B_L_No__c, Freight__c,
           Departure_Date_ETD__c, Arrival_Schedule_ETA__c,
           Remark_number_on_documents__c,
           Terms_of_Sales__c, Terms_of_Payment__c,
           Subtotal_USD__c, Fumigation__c, In_words__c,
           Total_Price_USD__c, Surcharge_amount_USD__c,
           Discount_Percentage__c, Discount_Amount__c
    FROM Shipment__c
    WHERE Id = '{shipment_id}'
    """
    shipment_result = sf.query(shipment_query)
    if not shipment_result["records"]:
        raise HTTPException(status_code=404, detail=f"No Shipment found with ID: {shipment_id}")
    shipment = shipment_result["records"][0]
    
    # Determine if discount exists
    discount_percentage = shipment.get("Discount_Percentage__c")
    discount_amount = shipment.get("Discount_Amount__c")
    discount_exists = any(
        v not in (None, 0, "", "0", 0.0)
        for v in (discount_percentage, discount_amount)
    )
    invoice_template_path = discount_invoice_template_path if discount_exists else base_invoice_template_path
    
    # Query account/consignee data
    if shipment.get("Consignee__c"):
        account_query = f"""
        SELECT Name, BillingStreet, BillingCity, BillingPostalCode, BillingCountry,
               Phone, Fax__c, VAT__c
        FROM Account
        WHERE Id = '{shipment['Consignee__c']}'
        """
        account_result = sf.query(account_query)
        account = account_result["records"][0] if account_result["records"] else {}
    else:
        account = {}
    
    # Query bookings (for packing list)
    bookings_query = f"""
    SELECT Id, Cont_Quantity__c
    FROM Booking__c
    WHERE Shipment__c = '{shipment_id}'
    """
    bookings_result = sf.query_all(bookings_query)
    bookings = bookings_result['records']
    total_containers_from_bookings = sum(booking.get('Cont_Quantity__c') or 0 for booking in bookings)
    
    # Query container items (both packing list and invoice use this)
    items_query = f"""
    SELECT Line_item_no_for_print__c, Product_Description__c,
           Length__c, Width__c, Height__c,
           Quantity_For_print__c, Unit_for_print__c,
           Crates__c, Packing__c, Order_No__c,
           Sales_Price_USD__c, Charge_Unit__c, Total_Price_USD__c,
           Container__r.Name, Container__r.Container_Weight_Regulation__c,
           Container__r.STT_Cont__c
    FROM Container_Item__c
    WHERE Shipment__c = '{shipment_id}'
    ORDER BY Line_item_no_for_print__c
    """
    items_result = sf.query_all(items_query)
    items = items_result["records"]
    
    # Query deposits (for invoice)
    deposit_query = f"""
    SELECT Contract_PI__r.Name, Reconciled_Amount__c
    FROM Receipt_Reconciliation__c
    WHERE Invoice__c = '{shipment_id}'
    """
    deposits = sf.query_all(deposit_query)["records"]
    
    # Query refunds (for invoice)
    refunds_query = f"""
    SELECT Reason, Refund_Amount__c
    FROM Case
    WHERE Refund_in_Shipment__c = '{shipment_id}'
    """
    refunds = sf.query_all(refunds_query)["records"]
    
    # ===== GENERATE PACKING LIST SHEET =====
    wb_packing = openpyxl.load_workbook(packing_list_template_path)
    ws_packing = wb_packing['PackingList']
    
    # Packing list replacements
    packing_replacements = {
        '{{Shipment__c.Consignee__r.Name}}': account.get('Name') or '',
        '{{Shipment__c.Consignee__r.BillingStreet}}': account.get('BillingStreet') or '',
        '{{Shipment__c.Consignee__r.BillingCity}}': account.get('BillingCity') or '',
        '{{Shipment__c.Consignee__r.BillingPostalCode}}': account.get('BillingPostalCode') or '',
        '{{Shipment__c.Consignee__r.BillingCountry}}': account.get('BillingCountry') or '',
        '{{Shipment__c.Consignee__r.Phone}}': account.get('Phone') or '',
        '{{Shipment__c.Consignee__r.Fax__c}}': account.get('Fax__c') or '',
        '{{Shipment__c.Consignee__r.VAT__c}}': account.get('VAT__c') or '',
        '{{Shipment__c.Invoice_Packing_list_no__c}}': shipment.get('Invoice_Packing_list_no__c') or '',
        '{{Shipment__c.Issued_date__c}}': shipment.get('Issued_date__c') or '',
        '{{Shipment__c.Port_of_Origin__c}}': shipment.get('Port_of_Origin__c') or '',
        '{{Shipment__c.Final_Destination__c}}': shipment.get('Final_Destination__c') or '',
        '{{Shipment__c.Stockyard__c}}': shipment.get('Stockyard__c') or '',
        '{{Shipment__c.Ocean_Vessel__c}}': shipment.get('Ocean_Vessel__c') or '',
        '{{Shipment__c.B_L_No__c}}': shipment.get('B_L_No__c') or '',
        '{{Shipment__c.Departure_Date_ETD__c}}': shipment.get('Departure_Date_ETD__c') or '',
        '{{Shipment__c.Arrival_Schedule_ETA__c}}': shipment.get('Arrival_Schedule_ETA__c') or '',
        '{{Shipment__c.Remark_number_on_documents__c}}': shipment.get('Remark_number_on_documents__c') or '',
        '{{Shipment__c.Terms_of_Sales__c}}': shipment.get('Terms_of_Sales__c') or '',
        '{{Shipment__c.Terms_of_Payment__c}}': shipment.get('Terms_of_Payment__c') or '',
    }
    
    for row in ws_packing.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for placeholder, value in packing_replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))
                if '{{TableStart:Shipment__c.r.Bookings__r}}' in cell.value:
                    cell.value = str(total_containers_from_bookings)
    
    # Remove "None" values
    for row in ws_packing.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'None' in cell.value:
                cell.value = cell.value.replace('None', '')
    
    # Handle freight checkboxes for packing list
    freight_value = (shipment.get('Freight__c') or '').strip()
    freight_upper = freight_value.upper()
    checked_box = '☑'
    unchecked_box = '☐'
    
    lines = []
    for opt in freight_options:
        mark = checked_box if opt.upper() == freight_upper else unchecked_box
        lines.append(f"{mark} {opt}")
    checkbox_text = "\n".join(lines)
    
    for row in ws_packing.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '{{Shipment__c.Freight__c}}' in cell.value:
                cell.value = cell.value.replace('{{Shipment__c.Freight__c}}', checkbox_text)
                if cell.alignment:
                    new_alignment = style_copy(cell.alignment)
                else:
                    from openpyxl.styles import Alignment
                    new_alignment = Alignment()
                new_alignment.wrap_text = True
                cell.alignment = new_alignment
    
    # Find table start row for packing list
    table_start_row = None
    for row in ws_packing.iter_rows(min_row=1, max_row=ws_packing.max_row, min_col=1, max_col=13):
        for cell in row:
            if cell.value and '{{TableStart:ContainerItems}}' in str(cell.value):
                table_start_row = cell.row
                break
        if table_start_row:
            break
    
    if not table_start_row:
        raise ValueError("No table start marker found in packing list template")
    
    # Expand table for packing list
    expand_items_table(ws_packing, table_start_row, len(items) if items else 1)
    
    # Fill in item data for packing list
    for idx, item in enumerate(items):
        row = table_start_row + idx
        container_r = item.get('Container__r', {})
        line_item_no = item.get('Line_item_no_for_print__c') or str(idx + 1)
        ws_packing.cell(row, 1).value = line_item_no
        ws_packing.cell(row, 2).value = item.get('Product_Description__c')
        ws_packing.cell(row, 3).value = item.get('Length__c')
        ws_packing.cell(row, 4).value = item.get('Width__c')
        ws_packing.cell(row, 5).value = item.get('Height__c')
        ws_packing.cell(row, 6).value = item.get('Quantity_For_print__c') or ''
        ws_packing.cell(row, 7).value = item.get('Unit_for_print__c') or ''
        ws_packing.cell(row, 8).value = item.get('Crates__c')
        ws_packing.cell(row, 9).value = f"{item.get('Packing__c') or ''} pcs/crate"
        ws_packing.cell(row, 10).value = container_r.get('Container_Weight_Regulation__c')
        ws_packing.cell(row, 11).value = container_r.get('Name')
        ws_packing.cell(row, 13).value = item.get('Order_No__c')
    
    # ===== GENERATE INVOICE SHEET =====
    wb_invoice = openpyxl.load_workbook(invoice_template_path)
    ws_invoice = wb_invoice["Invoice"] if "Invoice" in wb_invoice.sheetnames else wb_invoice.active
    
    # Format Port of Origin in uppercase
    port_of_origin = (shipment.get("Port_of_Origin__c") or "").upper()
    
    # Invoice replacements
    invoice_replacements = {
        "{{Shipment__c.Consignee__r.Name}}": account.get("Name") or "",
        "{{Shipment__c.Consignee__r.BillingStreet}}": account.get("BillingStreet") or "",
        "{{Shipment__c.Consignee__r.BillingCity}}": account.get("BillingCity") or "",
        "{{Shipment__c.Consignee__r.BillingPostalCode}}": account.get("BillingPostalCode") or "",
        "{{Shipment__c.Consignee__r.BillingCountry}}": account.get("BillingCountry") or "",
        "{{Shipment__c.Consignee__r.Phone}}": account.get("Phone") or "",
        "{{Shipment__c.Consignee__r.Fax__c}}": account.get("Fax__c") or "",
        "{{Shipment__c.Consignee__r.VAT__c}}": account.get("VAT__c") or "",
        "{{Shipment__c.Invoice_Packing_list_no__c}}": shipment.get("Invoice_Packing_list_no__c") or "",
        "{{Shipment__c.Issued_date__c}}": shipment.get("Issued_date__c") or "",
        "{{Shipment__c.Port_of_Origin__c}}": port_of_origin,
        "{{Shipment__c.Final_Destination__c}}": shipment.get("Final_Destination__c") or "",
        "{{Shipment__c.Stockyard__c}}": shipment.get("Stockyard__c") or "",
        "{{Shipment__c.Ocean_Vessel__c}}": shipment.get("Ocean_Vessel__c") or "",
        "{{Shipment__c.B_L_No__c}}": shipment.get("B_L_No__c") or "",
        "{{Shipment__c.Departure_Date_ETD__c}}": shipment.get("Departure_Date_ETD__c") or "",
        "{{Shipment__c.Arrival_Schedule_ETA__c}}": shipment.get("Arrival_Schedule_ETA__c") or "",
        "{{Shipment__c.Remark_number_on_documents__c}}": shipment.get("Remark_number_on_documents__c") or "",
        "{{Shipment__c.Subtotal_USD__c\\# #,##0.##}}": shipment.get("Subtotal_USD__c") or 0,
        "{{Shipment__c.Fumigation__c}}": shipment.get("Fumigation__c") or "",
        "{{Shipment__c.Total_Price_USD__c\\# #,##0.##}}": shipment.get("Total_Price_USD__c") or 0,
        "{{Shipment__c.In_words__c}}": shipment.get("In_words__c") or "",
        "{{Shipment__c.Discount_Percentage__c}}": shipment.get("Discount_Percentage__c") or "",
        "{{Shipment__c.Discount_Amount__c\\# #,##0.##}}": shipment.get("Discount_Amount__c") or 0,
    }
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for placeholder, value in invoice_replacements.items():
                    cell.value = cell.value.replace(placeholder, str(value))
    
    # Clean "None"
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "None" in cell.value:
                cell.value = cell.value.replace("None", "")
    
    # Format picklist fields with checkboxes (uppercase)
    freight_checkbox_text = format_picklist_checkboxes(
        freight_options, shipment.get("Freight__c"), uppercase=True
    )
    terms_of_sales_checkbox_text = format_picklist_checkboxes(
        terms_of_sales_options, shipment.get("Terms_of_Sales__c"), uppercase=True
    )
    terms_of_payment_checkbox_text = format_picklist_checkboxes(
        terms_of_payment_options, shipment.get("Terms_of_Payment__c"), uppercase=True
    )
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Shipment__c.Freight__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Freight__c}}", freight_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Sales__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Sales__c}}", terms_of_sales_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                if "{{Shipment__c.Terms_of_Payment__c}}" in cell.value:
                    cell.value = cell.value.replace("{{Shipment__c.Terms_of_Payment__c}}", terms_of_payment_checkbox_text)
                    cell.alignment = cell.alignment.copy(wrap_text=True)
    
    # Find ContainerItems table for invoice
    invoice_table_start_row = None
    for row in ws_invoice.iter_rows(min_row=1, max_row=ws_invoice.max_row, min_col=1, max_col=ws_invoice.max_column):
        for cell in row:
            if isinstance(cell.value, str) and "{{TableStart:ContainerItems}}" in cell.value:
                invoice_table_start_row = cell.row
                break
        if invoice_table_start_row:
            break
    
    if not invoice_table_start_row:
        raise ValueError("No ContainerItems table start marker found in invoice template")
    
    expand_invoice_items_table(ws_invoice, invoice_table_start_row, len(items) if items else 1)
    
    for idx, item in enumerate(items):
        row_idx = invoice_table_start_row + idx
        container_r = item.get("Container__r") or {}
        line_item_no = item.get("Line_item_no_for_print__c") or str(idx + 1)
        ws_invoice.cell(row_idx, 1).value = line_item_no
        ws_invoice.cell(row_idx, 2).value = item.get("Product_Description__c")
        ws_invoice.cell(row_idx, 3).value = item.get("Length__c")
        ws_invoice.cell(row_idx, 4).value = item.get("Width__c")
        ws_invoice.cell(row_idx, 5).value = item.get("Height__c")
        ws_invoice.cell(row_idx, 6).value = item.get("Quantity_For_print__c")
        ws_invoice.cell(row_idx, 7).value = item.get("Unit_for_print__c")
        ws_invoice.cell(row_idx, 8).value = container_r.get("STT_Cont__c") or container_r.get("Name")
        ws_invoice.cell(row_idx, 9).value = f"{item.get('Sales_Price_USD__c') or ''} {item.get('Charge_Unit__c') or ''}".strip()
        ws_invoice.cell(row_idx, 10).value = item.get("Total_Price_USD__c")
        ws_invoice.cell(row_idx, 11).value = item.get("Order_No__c")
    
    # Handle deposits / refunds / surcharge sections
    deposit_text_cell = None
    deposit_amount_cell = None
    refund_cell = None
    surcharge_text_cell = None
    surcharge_amount_cell = None
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            val = cell.value
            if "{{TableStart:InvoiceDeposit}}" in val:
                deposit_text_cell = cell
            if "Reconciled_Amount__c" in val:
                deposit_amount_cell = cell
            if "{{TableStart:Shipment__c.r.Cases__r}}" in val:
                refund_cell = cell
            if "{{TableStart:Surcharges}}" in val:
                surcharge_text_cell = cell
            if "Surcharge_amount_USD__c" in val:
                surcharge_amount_cell = cell
    
    if deposit_text_cell and deposit_amount_cell:
        if deposits:
            labels = []
            amounts = []
            for rec in deposits:
                pi_name = (rec.get("Contract_PI__r") or {}).get("Name") or ""
                labels.append(f"Deduct: Deposit of PI {pi_name}".strip())
                amt = rec.get("Reconciled_Amount__c")
                amounts.append("" if amt is None else f"{amt:,.2f}")
            deposit_text_cell.value = "\n".join(labels)
            deposit_amount_cell.value = "\n".join(amounts)
        else:
            deposit_text_cell.value = None
            deposit_amount_cell.value = None
    
    if refund_cell:
        if refunds:
            lines = []
            for rec in refunds:
                reason = rec.get("Reason") or ""
                amt = rec.get("Refund_Amount__c")
                part = reason
                if amt is not None:
                    part = f"{reason} {amt:,.2f}".strip()
                lines.append(part)
            refund_cell.value = "\n".join(lines)
        else:
            refund_cell.value = None
    
    for row in ws_invoice.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if "{{Refund_Amount__c" in cell.value or "{{TableEnd:Shipment__c.r.Cases__r}}" in cell.value:
                    if not refunds:
                        cell.value = cell.value.replace("{{Refund_Amount__c\\# #,##0.##}}", "")
                        cell.value = cell.value.replace("{{TableEnd:Shipment__c.r.Cases__r}}", "")
                        if cell.value and not cell.value.strip():
                            cell.value = None
    
    surcharge_amount = shipment.get("Surcharge_amount_USD__c")
    if surcharge_text_cell or surcharge_amount_cell:
        if surcharge_amount:
            if surcharge_text_cell:
                surcharge_text_cell.value = "Surcharge:"
            if surcharge_amount_cell:
                surcharge_amount_cell.value = f"{surcharge_amount:,.2f}"
        else:
            if surcharge_text_cell:
                surcharge_text_cell.value = None
            if surcharge_amount_cell:
                surcharge_amount_cell.value = None
    
    # ===== COMBINE INTO ONE WORKBOOK =====
    # Create a new workbook and copy sheets
    combined_wb = openpyxl.Workbook()
    combined_wb.remove(combined_wb.active)  # Remove default sheet
    
    # Copy packing list sheet
    ws_packing_copy = combined_wb.create_sheet("Packing List")
    for row in ws_packing.iter_rows():
        for cell in row:
            new_cell = ws_packing_copy.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = style_copy(cell.font)
                new_cell.border = style_copy(cell.border)
                new_cell.fill = style_copy(cell.fill)
                new_cell.number_format = style_copy(cell.number_format)
                new_cell.protection = style_copy(cell.protection)
                new_cell.alignment = style_copy(cell.alignment)
    
    # Copy column dimensions
    for col in ws_packing.column_dimensions:
        if col in ws_packing.column_dimensions:
            combined_wb["Packing List"].column_dimensions[col].width = ws_packing.column_dimensions[col].width
    
    # Copy row dimensions
    for row in ws_packing.row_dimensions:
        if row in ws_packing.row_dimensions:
            combined_wb["Packing List"].row_dimensions[row].height = ws_packing.row_dimensions[row].height
    
    # Copy merged cells
    for merged_cell in ws_packing.merged_cells.ranges:
        combined_wb["Packing List"].merge_cells(str(merged_cell))
    
    # Copy invoice sheet
    ws_invoice_copy = combined_wb.create_sheet("Invoice")
    for row in ws_invoice.iter_rows():
        for cell in row:
            new_cell = ws_invoice_copy.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = style_copy(cell.font)
                new_cell.border = style_copy(cell.border)
                new_cell.fill = style_copy(cell.fill)
                new_cell.number_format = style_copy(cell.number_format)
                new_cell.protection = style_copy(cell.protection)
                new_cell.alignment = style_copy(cell.alignment)
    
    # Copy column dimensions
    for col in ws_invoice.column_dimensions:
        if col in ws_invoice.column_dimensions:
            combined_wb["Invoice"].column_dimensions[col].width = ws_invoice.column_dimensions[col].width
    
    # Copy row dimensions
    for row in ws_invoice.row_dimensions:
        if row in ws_invoice.row_dimensions:
            combined_wb["Invoice"].row_dimensions[row].height = ws_invoice.row_dimensions[row].height
    
    # Copy merged cells
    for merged_cell in ws_invoice.merged_cells.ranges:
        combined_wb["Invoice"].merge_cells(str(merged_cell))
    
    
    # Save combined file
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"Combined_Export_{shipment.get('Invoice_Packing_list_no__c', shipment['Name'])}_{timestamp}.xlsx"
    
    # Use appropriate output directory based on environment
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    
    combined_wb.save(str(file_path))
    
    # Upload to Salesforce as ContentVersion
    with open(file_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode("utf-8")
    
    content_version = sf.ContentVersion.create(
        {
            "Title": file_name.rsplit(".", 1)[0],
            "PathOnClient": file_name,
            "VersionData": encoded,
            "FirstPublishLocationId": shipment_id,
        }
    )
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"],
        "sheets": ["Packing List", "Invoice"],
        "item_count": len(items),
        "deposit_count": len(deposits),
        "refund_count": len(refunds),
        "discount_exists": discount_exists,
        "template_used": {
            "packing_list": packing_list_template_path,
            "invoice": invoice_template_path
        }
    }

@app.get("/download/{file_name}")
async def download_file(file_name: str):
    """
    Download a generated packing list file
    
    Parameters:
    - file_name: Name of the file to download
    
    Note: In serverless environments (Vercel), files in /tmp are ephemeral.
    The download endpoint may not work reliably. Files are always uploaded to Salesforce.
    """
    # Try to find the file in the appropriate output directory
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    
    # Also check the legacy output directory for backwards compatibility
    if not file_path.exists():
        legacy_path = Path("output") / file_name
        if legacy_path.exists():
            file_path = legacy_path
    
    if not file_path.exists():
        raise HTTPException(
            status_code=404, 
            detail=f"File not found. In serverless environments, use the Salesforce attachment instead."
        )
    
    return FileResponse(
        path=str(file_path),
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- New Helper Functions for PI, PO, Quote ---

def expand_table_by_tag(ws, start_tag, end_tag, data):
    """
    Expand a single row table based on start and end tags.
    """
    # Find the row containing the tags
    table_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if start_tag in cell.value:
                    table_row_idx = cell.row
                    break
        if table_row_idx:
            break
            
    if not table_row_idx:
        print(f"Warning: Table tags {start_tag} not found.")
        return None

    if not data:
        # Clear tags and placeholders, keep static text
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=table_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                # Remove tags
                val = cell.value.replace(start_tag, "").replace(end_tag, "")
                # Remove any remaining placeholders {{...}}
                val = re.sub(r"\{\{.*?\}\}", "", val)
                cell.value = val
        return table_row_idx

    num_rows = len(data)
    add_rows = max(0, num_rows - 1)
    
    # Capture styles from the template row
    max_col = ws.max_column
    row_style = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=table_row_idx, column=col)
        row_style.append(style_copy(cell._style) if cell.has_style else None)
    
    row_height = ws.row_dimensions[table_row_idx].height

    # Handle merged cells (shift them down)
    merges_to_shift = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > table_row_idx:
            merges_to_shift.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col))
    
    for mr in merges_to_shift:
        rng = f"{get_column_letter(mr[2])}{mr[0]}:{get_column_letter(mr[3])}{mr[1]}"
        ws.unmerge_cells(rng)

    # Insert rows if needed
    if add_rows > 0:
        ws.insert_rows(table_row_idx + 1, amount=add_rows)
        
        for offset in range(1, add_rows + 1):
            r = table_row_idx + offset
            # Copy row height
            if row_height is not None:
                ws.row_dimensions[r].height = row_height
                
            for col in range(1, max_col + 1):
                dst = ws.cell(row=r, column=col)
                # Copy value from template row (to preserve placeholders)
                src_val = ws.cell(row=table_row_idx, column=col).value
                dst.value = src_val
                
                # Copy style
                st = row_style[col - 1]
                if st is not None:
                    dst._style = style_copy(st)
                    
    # Re-merge shifted cells
    for mr in merges_to_shift:
        new_min_row = mr[0] + add_rows
        new_max_row = mr[1] + add_rows
        rng = f"{get_column_letter(mr[2])}{new_min_row}:{get_column_letter(mr[3])}{new_max_row}"
        ws.merge_cells(rng)
                    
    # Fill data
    for i, record in enumerate(data):
        current_row_idx = table_row_idx + i
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=current_row_idx, column=col)
            if cell.value and isinstance(cell.value, str):
                cell_val = cell.value.replace(start_tag, "").replace(end_tag, "")
                
                for key, value in record.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in cell_val:
                        cell_val = cell_val.replace(placeholder, str(value) if value is not None else "")
                        
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, cell_val)
                    for fmt in matches:
                        try:
                            if value is not None:
                                if isinstance(value, (int, float)):
                                    if "#,##0.##" in fmt:
                                         formatted_val = "{:,.2f}".format(value)
                                    else:
                                         formatted_val = str(value)
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                                else:
                                    cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))
                            else:
                                cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", "")
                        except:
                             cell_val = cell_val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value))

                cell.value = cell_val
                
                # Attempt to convert to number if it looks like one
                if isinstance(cell.value, str):
                    try:
                        clean_val = cell.value.replace(',', '')
                        f_val = float(clean_val)
                        is_leading_zero = (len(clean_val) > 1 and clean_val.startswith('0') and not clean_val.startswith('0.'))
                        
                        if not is_leading_zero:
                            if f_val.is_integer():
                                cell.value = int(f_val)
                            else:
                                cell.value = f_val
                    except ValueError:
                        pass
    
    return table_row_idx

# --- PI No Discount Generation ---

def generate_pi_no_discount_file(contract_id: str, template_path: str):
    sf = get_salesforce_connection()
    
    # Query Contract
    contract_fields = [
        "Id", "Name", "CreatedDate", "Port_of_Origin__c", "Port_of_Discharge__c", "Stockyard__c",
        "Account__r.Name", "Account__r.BillingStreet", "Account__r.BillingCity", 
        "Account__r.BillingPostalCode", "Account__r.BillingCountry", "Account__r.Phone", 
        "Account__r.Fax__c", "Account__r.VAT__c",
        "Export_Route_Carrier__c", "Incoterms__c", "Packing__c", "Shipping_Schedule__c",
        "Terms_of_Sale__c", "Terms_of_Payment__c", "REMARK_NUMBER_ON_DOCUMENTS__c",
        "Total_Crates__c", "Total_m3__c", "Total_Tons__c", "Total_Conts__c",
        "Sub_Total_USD__c", "In_words__c", "Fumigation__c", "Discount__c", 
        "Discount_Amount__c", "Deposit_Percentage__c", "Deposit__c", "Total_Price_USD__c",
        "Created_Date__c"
    ]
    
    query = f"SELECT {', '.join(contract_fields)} FROM Contract__c WHERE Id = '{contract_id}'"
    try:
        result = sf.query(query)
    except Exception as e:
        # Fallback query if some fields don't exist
        print(f"Error querying contract: {e}")
        raise ValueError(f"Error querying contract: {e}")

    if not result['records']:
        raise ValueError(f"No contract found with ID: {contract_id}")

    contract = result['records'][0]
    
    # Flatten Account__r fields
    if contract.get('Account__r'):
        for k, v in contract['Account__r'].items():
            contract[f'Account__r.{k}'] = v
            
    full_data = {}
    for k, v in contract.items():
        full_data[f"Contract__c.{k}"] = v
        
    # Query Products
    product_fields = [
        "Line_number_For_print__c", "Product_Discription__c", "L_PI__c", "W_PI__c", "H_PI__c",
        "PCS_PI__c", "m2__c", "Crates_PI__c", "m3__c", "Tons__c", "Cont__c",
        "Sales_Price__c", "Charge_Unit__c", "Total_Price_USD__c", "Packing_PI__c",
        "Product__r.Name"
    ]
    prod_query = f"SELECT {', '.join(product_fields)} FROM Contract_Product__c WHERE Contract__r.Id = '{contract_id}' ORDER BY Line_Number__c ASC"
    try:
        prod_result = sf.query_all(prod_query)
        contract_items = prod_result['records']
    except Exception as e:
        print(f"Error querying products: {e}")
        contract_items = []

    # Inject Sequential Number
    for idx, item in enumerate(contract_items):
        item['Line_number_For_print__c'] = idx + 1

    # Query Surcharges
    surcharge_query = f"SELECT Id, Name, Surcharge_amount_USD__c FROM Expense__c WHERE Contract_PI__r.Id = '{contract_id}' AND Surcharge_amount_USD__c != 0"
    try:
        sur_result = sf.query_all(surcharge_query)
        surcharge_records = sur_result['records']
    except Exception as e:
        surcharge_records = []
        
    surcharge_items = []
    for item in surcharge_records:
        surcharge_items.append({
            "Name": item.get('Name'),
            "Surcharge_amount_USD__c": item.get('Surcharge_amount_USD__c')
        })

    # Load Template
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill Main Data
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                
                # Conditional Logic
                if_pattern = r"\{\{#if\s+([\w\.]+)\s+'=='\s+'([^']+)'\}\}(.*?)\{\{else\}\}(.*?)\{\{/if\}\}"
                if_matches = re.findall(if_pattern, val)
                for match in if_matches:
                    key, target_val, true_text, false_text = match
                    full_match_str = f"{{{{#if {key} '==' '{target_val}'}}}}{true_text}{{{{else}}}}{false_text}{{{{/if}}}}"
                    actual_val = str(full_data.get(key, ""))
                    if actual_val.lower() == target_val.lower():
                        val = val.replace(full_match_str, true_text)
                    else:
                        val = val.replace(full_match_str, false_text)

                # Float Fields
                float_fields = [
                    "{{Contract__c.Total_Crates__c}}", "{{Contract__c.Total_m3__c}}",
                    "{{Contract__c.Total_Tons__c}}", "{{Contract__c.Total_Conts__c}}",
                    "{{Contract__c.Sub_Total_USD__c\\# #,##0.##}}",
                    "{{Contract__c.Total_Price_USD__c\\# #,##0.##}}",
                    "{{Contract__c.Deposit__c\\# #,##0.##}}"
                ]
                is_float_field = False
                for field in float_fields:
                    if field in val:
                        key_part = field.replace("{{", "").replace("}}", "").split("\\#")[0]
                        value = full_data.get(key_part)
                        if value is not None:
                            try:
                                cell.value = float(value)
                                cell.number_format = '#,##0.00'
                                is_float_field = True
                            except ValueError:
                                pass
                        break
                if is_float_field:
                    continue

                # General Replacement
                for key, value in full_data.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in val:
                        val = val.replace(placeholder, str(value) if value is not None else "")
                    
                    pattern = f"\\{{{{{key}\\\\#(.*?)\\}}}}"
                    matches = re.findall(pattern, val)
                    for fmt in matches:
                         if value is not None and isinstance(value, (int, float)):
                             if "#,##0.##" in fmt:
                                 formatted_val = "{:,.2f}".format(value)
                             else:
                                 formatted_val = str(value)
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", formatted_val)
                         else:
                             val = val.replace(f"{{{{{key}\\#{fmt}}}}}", str(value) if value is not None else "")
                cell.value = val

    # Fill Product Table
    table_start_row = expand_table_by_tag(ws, "{{TableStart:ContractProduct2}}", "{{TableEnd:ContractProduct2}}", contract_items)
    
    # Bold Logic & Merging (Simplified from script)
    if table_start_row and contract_items:
        col_b_idx = 2
        for i, item in enumerate(contract_items):
            row_idx = table_start_row + i
            cell = ws.cell(row=row_idx, column=col_b_idx)
            product_name = item.get('Product__r', {}).get('Name')
            current_desc = str(cell.value) if cell.value else ""
            
            if product_name and current_desc:
                match = re.match(r"^([^\d\(]+)", product_name)
                if match:
                    bold_target = match.group(1).strip()
                    if bold_target and bold_target in current_desc:
                        start_idx = current_desc.find(bold_target)
                        if start_idx != -1:
                            parts = []
                            if start_idx > 0: parts.append(current_desc[:start_idx])
                            parts.append(TextBlock(InlineFont(b=True), bold_target))
                            end_idx = start_idx + len(bold_target)
                            if end_idx < len(current_desc): parts.append(current_desc[end_idx:])
                            cell.value = CellRichText(parts)

        # Merge identical cells
        start_merge_row = table_start_row
        current_val = str(ws.cell(row=start_merge_row, column=col_b_idx).value)
        for i in range(1, len(contract_items)):
            row_idx = table_start_row + i
            cell_val = str(ws.cell(row=row_idx, column=col_b_idx).value)
            if cell_val != current_val:
                if row_idx - 1 > start_merge_row:
                    ws.merge_cells(start_row=start_merge_row, start_column=col_b_idx, end_row=row_idx-1, end_column=col_b_idx)
                    ws.cell(row=start_merge_row, column=col_b_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                start_merge_row = row_idx
                current_val = cell_val
        # Last block
        last_row = table_start_row + len(contract_items) - 1
        if last_row > start_merge_row:
            ws.merge_cells(start_row=start_merge_row, start_column=col_b_idx, end_row=last_row, end_column=col_b_idx)
            ws.cell(row=start_merge_row, column=col_b_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Fill Surcharges
    expand_table_by_tag(ws, "{{TableStart:PISurcharge}}", "{{TableEnd:PISurcharge}}", surcharge_items)

    # Save
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"PI_NoDiscount_{contract.get('Name')}_{timestamp}.xlsx"
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": contract_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }

@app.get("/generate-pi-no-discount/{contract_id}")
async def generate_pi_no_discount_endpoint(contract_id: str):
    try:
        template_path = os.getenv('PI_NO_DISCOUNT_TEMPLATE_PATH', 'templates/proforma_invoice_template_no_discount.xlsx')
        if not os.path.exists(template_path):
             # Fallback to root if not in templates
             if os.path.exists('proforma_invoice_template_no_discount.xlsx'):
                 template_path = 'proforma_invoice_template_no_discount.xlsx'
             else:
                 raise HTTPException(status_code=404, detail=f"Template not found: {template_path}")
        
        result = generate_pi_no_discount_file(contract_id, template_path)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Production Order Generation ---

def generate_production_order_file(contract_id: str, template_path: str):
    sf = get_salesforce_connection()
    
    # Query Contract
    contract_query = f"""
        SELECT Id, Production_Order_Number__c, Name, CreatedDate, Port_of_Origin__c, 
               Port_of_Discharge__c, Stockyard__c, Total_Pcs_PO__c, Total_Crates__c, 
               Total_m2__c, Total_m3__c, Total_Tons__c, Total_Conts__c
        FROM Contract__c 
        WHERE Id = '{contract_id}'
    """
    contract_res = sf.query(contract_query)
    if not contract_res['records']:
        raise ValueError(f"Contract not found: {contract_id}")
    contract_data = contract_res['records'][0]

    # Query Products
    products_query = f"""
        SELECT Id, Name, Charge_Unit__c, Cont__c, Crates__c, Height__c, Length__c, Quantity__c, Width__c, m2__c, m3__c, Packing__c, Tons__c, Product_Description__c, Delivery_Date__c, Line_number__c, SKU__c, Vietnamese_Description__c, Order__r.Name 
        FROM Order_Product__c 
        WHERE Contract_PI__r.Id = '{contract_id}' 
        ORDER BY Line_number__c ASC
    """
    products_res = sf.query(products_query)
    products_data = products_res['records']

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Flatten data
    flat_data = {}
    for k, v in contract_data.items():
        flat_data[f"Contract__c.{k}"] = v
        if "Date" in k and v:
            try:
                dt = datetime.datetime.strptime(v[:10], "%Y-%m-%d")
                flat_data[f"Contract__c.{k}\\@dd/MM/yyyy"] = dt.strftime("%d/%m/%Y")
            except: pass

    # Fill placeholders
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                matches = re.findall(r"\{\{([^\}]+)\}\}", val)
                for match in matches:
                    key_part = match.split('\\')[0].strip()
                    format_part = match.split('\\@')[1].strip() if '\\@' in match else None
                    
                    if key_part in flat_data:
                        replace_val = flat_data[key_part]
                        if replace_val is None: replace_val = ""
                        
                        if format_part and replace_val:
                            try:
                                val_str = str(replace_val).split('T')[0]
                                dt = datetime.datetime.strptime(val_str, "%Y-%m-%d")
                                py_format = format_part.replace('dd', '%d').replace('MM', '%m').replace('yyyy', '%Y')
                                replace_val = dt.strftime(py_format)
                            except: pass
                        
                        val = val.replace(f"{{{{{match}}}}}", str(replace_val))
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal=cell.alignment.horizontal if cell.alignment else 'left')
                cell.value = val

    # Fill Table
    table_start_row = None
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val and "{{TableStart:ProPlanProduct}}" in str(cell_val):
            table_start_row = r
            break
            
    if table_start_row and products_data:
        num_items = len(products_data)
        if num_items > 1:
            ws.insert_rows(table_start_row + 1, amount=num_items - 1)
            
        # Copy styles (simplified)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Copy styles from template row
        if num_items > 1:
            for i in range(1, num_items):
                for col in range(1, 16):
                    source_cell = ws.cell(row=table_start_row, column=col)
                    target_cell = ws.cell(row=table_start_row + i, column=col)
                    if source_cell.border: target_cell.border = style_copy(source_cell.border)
                    if source_cell.font: target_cell.font = style_copy(source_cell.font)
                    if source_cell.alignment: target_cell.alignment = style_copy(source_cell.alignment)
                    if source_cell.number_format: target_cell.number_format = style_copy(source_cell.number_format)

        ws.cell(row=table_start_row, column=1).value = ""
        
        for i, item in enumerate(products_data):
            row_idx = table_start_row + i
            
            # Unmerge if needed
            for col in range(1, 16):
                cell = ws.cell(row=row_idx, column=col)
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        try: ws.unmerge_cells(str(merged_range))
                        except: pass
                ws.cell(row=row_idx, column=col).border = thin_border

            # Map Data
            ws.cell(row=row_idx, column=1).value = i + 1
            ws.cell(row=row_idx, column=1).alignment = align_center
            ws.cell(row=row_idx, column=2).value = item.get("Order__r", {}).get("Name") if item.get("Order__r") else ""
            ws.cell(row=row_idx, column=2).alignment = align_center
            ws.cell(row=row_idx, column=3).value = item.get("SKU__c")
            ws.cell(row=row_idx, column=3).alignment = align_left
            
            # Rich Text Description
            desc_val = item.get("Vietnamese_Description__c") or ""
            if desc_val and '-' in str(desc_val):
                parts = str(desc_val).split('-', 1)
                rich_text = CellRichText(
                    TextBlock(InlineFont(b=True, rFont='Times New Roman', sz=11), parts[0]),
                    TextBlock(InlineFont(b=False, rFont='Times New Roman', sz=11), '-' + parts[1])
                )
                ws.cell(row=row_idx, column=4).value = rich_text
            else:
                ws.cell(row=row_idx, column=4).value = desc_val
            ws.cell(row=row_idx, column=4).alignment = align_left
            
            # Dimensions & Quantity
            ws.cell(row=row_idx, column=5).value = item.get("Length__c")
            ws.cell(row=row_idx, column=6).value = item.get("Width__c")
            ws.cell(row=row_idx, column=7).value = item.get("Height__c")
            ws.cell(row=row_idx, column=8).value = item.get("Quantity__c")
            ws.cell(row=row_idx, column=9).value = item.get("Crates__c")
            
            if item.get("m2__c"): 
                ws.cell(row=row_idx, column=10).value = float(item.get("m2__c"))
                ws.cell(row=row_idx, column=10).number_format = '0.00'
            if item.get("m3__c"):
                ws.cell(row=row_idx, column=11).value = float(item.get("m3__c"))
                ws.cell(row=row_idx, column=11).number_format = '0.00'
                
            ws.cell(row=row_idx, column=12).value = item.get("Tons__c")
            ws.cell(row=row_idx, column=13).value = item.get("Cont__c")
            
            for col in range(5, 14): ws.cell(row=row_idx, column=col).alignment = align_center
            
            # Packing
            packing_val = item.get("Packing__c")
            if packing_val:
                try:
                    ws.cell(row=row_idx, column=14).value = float(packing_val)
                    ws.cell(row=row_idx, column=14).number_format = '0.0 "viên/kiện"'
                except:
                    ws.cell(row=row_idx, column=14).value = f"{packing_val}\nviên/kiện"
            ws.cell(row=row_idx, column=14).alignment = align_center
            
            # Delivery Date
            del_date = item.get("Delivery_Date__c")
            if del_date:
                try:
                    dt = datetime.datetime.strptime(del_date[:10], "%Y-%m-%d")
                    ws.cell(row=row_idx, column=15).value = dt.strftime("%d/%m/%Y")
                except:
                    ws.cell(row=row_idx, column=15).value = del_date
            ws.cell(row=row_idx, column=15).alignment = align_center

    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"ProductionOrder_{contract_data.get('Production_Order_Number__c', 'Draft')}_{timestamp}.xlsx"
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": contract_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }

@app.get("/generate-production-order/{contract_id}")
async def generate_production_order_endpoint(contract_id: str):
    try:
        template_path = os.getenv('PRODUCTION_ORDER_TEMPLATE_PATH', 'templates/production_order_template.xlsx')
        if not os.path.exists(template_path):
             if os.path.exists('production_order_template.xlsx'):
                 template_path = 'production_order_template.xlsx'
             else:
                 raise HTTPException(status_code=404, detail=f"Template not found: {template_path}")
        
        result = generate_production_order_file(contract_id, template_path)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Quote No Discount Generation ---

def generate_quote_no_discount_file(quote_id: str, template_path: str):
    sf = get_salesforce_connection()
    
    # Query Quote Items
    query = f"""
    SELECT Id, Quote.Name, Quote.AccountId, Quote.Total_Crates__c, Quote.Total_m3__c, 
           Quote.Total_Tons__c, Quote.Total_Conts__c, Quote.Sub_Total_USD__c, Quote.Total_Price_USD__c,
           Product_Name__c, Product_Description__c, Length__c, Width__c, Height__c, Quantity, 
           Crates__c, m2__c, m3__c, Tons__c, Cont__c, Packing__c, Unit_Price_USD__c, Total_Price_USD__c,
           Quote_Line_Item_Number_Quote__c
    FROM QuoteLineItem 
    WHERE QuoteId = '{quote_id}' 
    ORDER BY Quote_Line_Item_Number_Quote__c ASC
    """
    result = sf.query_all(query)
    if not result['records']:
        # Try fetching just Quote
        q_res = sf.query(f"SELECT Id, Name, AccountId FROM Quote WHERE Id = '{quote_id}'")
        if not q_res['records']: raise ValueError("Quote not found")
        quote_data = q_res['records'][0]
        quote_items = []
    else:
        quote_items = result['records']
        quote_data = quote_items[0]['Quote']

    # Flatten Data
    full_data = {}
    for k, v in quote_data.items():
        full_data[f"Quote.{k}"] = v
        
    if quote_data.get('AccountId'):
        try:
            acc = sf.Account.get(quote_data['AccountId'])
            full_data[f"Quote.Account.Name"] = acc.get('Name')
            full_data[f"Quote.Account.BillingStreet"] = acc.get('BillingStreet')
            # ... add other fields as needed
        except: pass

    # Inject Sequential Number
    for idx, item in enumerate(quote_items):
        item['Quote_Line_Item_Number_Quote__c'] = idx + 1

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill Main Data
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                # Simple replacement for Quote
                for key, value in full_data.items():
                    placeholder = f"{{{{{key}}}}}"
                    if placeholder in val:
                        val = val.replace(placeholder, str(value) if value is not None else "")
                cell.value = val

    # Fill Product Table
    expand_table_by_tag(ws, "{{TableStart:GetQuoteLine}}", "{{TableEnd:GetQuoteLine}}", quote_items)

    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"Quote_{quote_data.get('Name')}_{timestamp}.xlsx"
    output_dir = get_output_directory()
    file_path = output_dir / file_name
    wb.save(str(file_path))
    
    # Upload to Salesforce
    with open(file_path, "rb") as f:
        file_data = f.read()
    encoded = base64.b64encode(file_data).decode("utf-8")
    
    content_version = sf.ContentVersion.create({
        "Title": file_name.rsplit(".", 1)[0],
        "PathOnClient": file_name,
        "VersionData": encoded,
        "FirstPublishLocationId": quote_id
    })
    
    return {
        "file_path": str(file_path),
        "file_name": file_name,
        "salesforce_content_version_id": content_version["id"]
    }

@app.get("/generate-quote-no-discount/{quote_id}")
async def generate_quote_no_discount_endpoint(quote_id: str):
    try:
        template_path = os.getenv('QUOTE_NO_DISCOUNT_TEMPLATE_PATH', 'templates/quotation_template_no_discount.xlsx')
        if not os.path.exists(template_path):
             if os.path.exists('quotation_template_no_discount.xlsx'):
                 template_path = 'quotation_template_no_discount.xlsx'
             else:
                 raise HTTPException(status_code=404, detail=f"Template not found: {template_path}")
        
        result = generate_quote_no_discount_file(quote_id, template_path)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
